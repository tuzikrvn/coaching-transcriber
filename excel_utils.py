"""
Excel export utilities for coaching session transcriptions.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
SPEAKER_COLORS = ["D6E4F0", "D5F5E3", "FDEBD0", "E8DAEF"]
PAUSE_COLOR = "FEF9E7"
PAUSE_LONG_COLOR = "FCF3CF"  # pauses >= 5 sec
HEADER_COLOR = "1A237E"

SPEAKER_LABELS = ["А", "Б", "В", "Г"]


def format_time(ms: int) -> str:
    """Convert milliseconds to MM:SS or HH:MM:SS format."""
    total_sec = ms / 1000
    h = int(total_sec // 3600)
    m = int((total_sec % 3600) // 60)
    s = total_sec % 60
    if h > 0:
        return f"{h:02d}:{m:02d}:{s:05.2f}"
    return f"{m:02d}:{s:05.2f}"


def detect_pauses(utterances, threshold_sec: float = 2.0) -> list:
    """
    Process utterances and insert pause rows where gaps exceed the threshold.
    Returns list of row dicts ready for Excel export.
    """
    rows = []
    speaker_map = {}

    for i, utt in enumerate(utterances):
        # Map raw speaker ID ("A", "B", …) to human-readable label
        sp = utt.speaker
        if sp not in speaker_map:
            idx = len(speaker_map)
            label = SPEAKER_LABELS[idx] if idx < len(SPEAKER_LABELS) else sp
            speaker_map[sp] = f"Собеседник {label}"

        rows.append({
            "type": "speech",
            "start_ms": utt.start,
            "end_ms": utt.end,
            "start_fmt": format_time(utt.start),
            "end_fmt": format_time(utt.end),
            "duration": round((utt.end - utt.start) / 1000, 1),
            "speaker": speaker_map[sp],
            "speaker_key": sp,
            "text": utt.text,
        })

        # Check gap to next utterance
        if i < len(utterances) - 1:
            gap_ms = utterances[i + 1].start - utt.end
            gap_sec = gap_ms / 1000
            if gap_sec >= threshold_sec:
                rows.append({
                    "type": "pause",
                    "start_ms": utt.end,
                    "end_ms": utterances[i + 1].start,
                    "start_fmt": format_time(utt.end),
                    "end_fmt": format_time(utterances[i + 1].start),
                    "duration": round(gap_sec, 1),
                    "speaker": "",
                    "speaker_key": "",
                    "text": "⏸ ПАУЗА",
                })

    return rows


def create_excel(rows: list, output_path: str, source_filename: str = "") -> None:
    """Create formatted Excel workbook with transcription and summary sheets."""
    wb = Workbook()

    # ── Transcription sheet ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Транскрипция"

    headers = ["Начало", "Конец", "Длит., сек", "Говорящий", "Текст"]
    col_widths = [10, 10, 13, 22, 90]

    thin_border = Border(bottom=Side(style="thin", color="D5D8DC"))

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Build speaker → color map
    speaker_color_map = {}
    used = 0
    for row_data in rows:
        sp_key = row_data.get("speaker_key", "")
        if sp_key and sp_key not in speaker_color_map:
            speaker_color_map[sp_key] = SPEAKER_COLORS[used % len(SPEAKER_COLORS)]
            used += 1

    # Data rows
    for r_idx, row_data in enumerate(rows, 2):
        is_pause = row_data["type"] == "pause"
        is_long_pause = is_pause and row_data["duration"] >= 5

        if is_long_pause:
            bg = PAUSE_LONG_COLOR
            font_color, bold = "7D6608", True
        elif is_pause:
            bg = PAUSE_COLOR
            font_color, bold = "9A7D0A", True
        else:
            bg = speaker_color_map.get(row_data.get("speaker_key", ""), "FFFFFF")
            font_color, bold = "1C2833", False

        values = [
            row_data["start_fmt"],
            row_data["end_fmt"],
            row_data["duration"],
            row_data["speaker"],
            row_data["text"],
        ]

        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(color=font_color, bold=bold, size=10)
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c_idx == 5),
                horizontal="left" if c_idx == 5 else "center",
            )
            cell.border = thin_border

        ws.row_dimensions[r_idx].height = 14 if is_pause else 48

    # ── Summary sheet ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Сводка")
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 32

    speech_rows = [r for r in rows if r["type"] == "speech"]
    pause_rows = [r for r in rows if r["type"] == "pause"]
    total_dur = rows[-1]["end_ms"] / 1000 if rows else 0

    summary = [
        ("Файл", source_filename),
        ("Общая длительность", f"{int(total_dur // 60)} мин {int(total_dur % 60)} сек"),
        ("Всего реплик", len(speech_rows)),
        ("Всего пауз (≥ 2 сек)", len(pause_rows)),
        ("", ""),
    ]

    # Per-speaker time
    speaker_totals = {}
    for r in speech_rows:
        sp = r["speaker"]
        speaker_totals[sp] = speaker_totals.get(sp, 0) + r["duration"]

    for sp, dur in sorted(speaker_totals.items()):
        pct = (dur / total_dur * 100) if total_dur > 0 else 0
        summary.append((
            f"Время речи — {sp}",
            f"{int(dur // 60)} мин {int(dur % 60)} сек ({pct:.0f}%)"
        ))

    if pause_rows:
        avg_p = sum(r["duration"] for r in pause_rows) / len(pause_rows)
        max_p = max(r["duration"] for r in pause_rows)
        long_pauses = [r for r in pause_rows if r["duration"] >= 5]
        summary += [
            ("", ""),
            ("Средняя длина паузы", f"{avg_p:.1f} сек"),
            ("Максимальная пауза", f"{max_p:.1f} сек"),
            ("Длинных пауз (≥ 5 сек)", len(long_pauses)),
        ]

    for r_idx, (label, value) in enumerate(summary, 1):
        cell_a = ws2.cell(row=r_idx, column=1, value=label)
        cell_b = ws2.cell(row=r_idx, column=2, value=value)
        if label:
            cell_a.font = Font(bold=True, size=11)
            cell_b.font = Font(size=11)
        ws2.row_dimensions[r_idx].height = 22

    wb.save(output_path)
